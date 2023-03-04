#!/usr/bin/env python

import argparse, random, copy, xlwt, xlrd, math, os
from datetime import datetime, timedelta
import numpy as np


def generate_cvrptw_data(dataset_size, instance_size):
    """
    In our project, we only have 4 columns to generate: Arrival Time, Depature Time, Location, Type,
    + (ID, Flight_No, Delay Avg, Delay Var)
    """
    print("---- Generating data ...")
    # Arrival Time
    # arr_time = np.random.randint(0, 1400, size=(dataset_size, instance_size))
    # Empirically, [5, 25] flights per hour
    arr_time = []
    flight_per_hr = np.random.randint(5, 20, size=(dataset_size, 23)).tolist()
    # print(flight_per_hr)
    for i in range(dataset_size):
        # check whether the flights are enough
        if sum(flight_per_hr[i]) < instance_size:
            extra = math.ceil((instance_size - sum(flight_per_hr[i])) / len(flight_per_hr[i]))
            flight_per_hr[i] = [ele + extra for ele in flight_per_hr[i]]
        cur_t, remain_flight = -1, instance_size
        arr_time.append([])
        for j in flight_per_hr[i]:
            cur_t = cur_t + 1
            if remain_flight < j:
                arr_time[i] += np.random.randint(cur_t * 60, (cur_t + 1) * 60, size=remain_flight).tolist()
                break
            else:
                arr_time[i] += np.random.randint(cur_t * 60, (cur_t + 1) * 60, size=j).tolist()
                remain_flight -= j
    arr_time = np.array(arr_time)
    # print(arr_time)

    # Depature Time
    stay = np.random.randint(30, 40, size=(dataset_size, instance_size))
    dep_time = arr_time + stay

    # Location
    from projectSolver import Loadlocation
    locs, location = {}, []
    data = Loadlocation("./test_200304/testPrj.cfg")
    for i in data:
        if i["name"] != "depot":
            locs[i["name"]] = -1
    for i in range(dataset_size):
        all_loc = copy.deepcopy(locs)
        location.append([])
        for j in range(instance_size):
            ava_locs = [loc for loc in all_loc if all_loc[loc] < arr_time[i][j]]
            choose_loc = random.sample(ava_locs, 1)[0]
            location[i].append(choose_loc)
            all_loc[choose_loc] = dep_time[i][j]

    # Type
    types = np.random.randint(0, 3, size=(dataset_size, instance_size)) + 1
    # Delay Avg and Delay Var
    delay_avg = np.zeros((dataset_size, instance_size), dtype=np.int8)
    delay_var = np.ones((dataset_size, instance_size), dtype=np.int8)
    # ID and Flight_No
    ids = [list(range(1, instance_size+1)) for _ in range(dataset_size)]
    flight_ids = ids

    # convert time(int) to datetime object
    arr_time, dep_time = arr_time.tolist(), dep_time.tolist()
    arr_datetime, dep_datetime = [], []
    for i in range(dataset_size):
        arr_datetime.append([])
        dep_datetime.append([])
        for j in range(instance_size):
            hr, hr1 = arr_time[i][j] / 60, dep_time[i][j] / 60
            min, min1 = arr_time[i][j] % 60, dep_time[i][j] % 60
            sec = 0
            str_t, str_t1 = str(int(hr)) + ':' + str(int(min)) + ':' + str(int(sec)), str(int(hr1)) + ':' + str(int(min1)) + ':' + str(int(sec))
            arr_datetime[i].append(str_t)
            dep_datetime[i].append(str_t1)

    return {"ID": ids, "Flight No.": flight_ids, "Arrival Time": arr_datetime, "Departure Time": dep_datetime,
            "Location": location, "Type": types.tolist(), "Delay Avg": delay_avg.tolist(), "Delay Var": delay_var.tolist()}


def write2xlsx(data, filenames):
    print("---- Writing to Files ...")
    for i, path in enumerate(filenames):
        book = xlwt.Workbook()
        sh = book.add_sheet('Sheet1', cell_overwrite_ok=True)
        for col, key in enumerate(data):
            sh.write(0, col, key)
            for row, value in enumerate(data[key][i], start=1):
                sh.write(row, col, value)
        book.save(path)

        # conversion to xlsx
        from openpyxl.workbook import Workbook as openpyxlWorkbook
        xlsBook = xlrd.open_workbook(path)
        workbook = openpyxlWorkbook()
        for i in range(0, xlsBook.nsheets):
            xlsSheet = xlsBook.sheet_by_index(i)
            sheet = workbook.active if i == 0 else workbook.create_sheet()
            sheet.title = xlsSheet.name
            for row in range(0, xlsSheet.nrows):
                for col in range(0, xlsSheet.ncols):
                    if row > 0 and (col == 2 or col == 3):  # Arrival Time, Departure Time
                        stri = xlsSheet.cell_value(row, col)
                        datetime_object = datetime.strptime(stri, "%H:%M:%S")
                        sheet.cell(row=row + 1, column=col + 1).value = datetime_object - datetime(1900, 1, 1)
                    else:
                        sheet.cell(row=row + 1, column=col + 1).value = xlsSheet.cell_value(row, col)
        os.remove(path)
        workbook.save(path+"x")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--filename", default='schedule', help="Filename of the dataset to create (ignores datadir)")
    parser.add_argument("--data_dir", default='data', help="Create datasets in data_dir/problem (default 'data')")
    parser.add_argument("--problem", type=str, default='cvrptw',
                        help="Problem, 'cvrptw', etc.")
    parser.add_argument('--dataset_size', type=int, default=100, help="How many datasets you want to generate")
    parser.add_argument('--instance_sizes', type=int, nargs='+', default=[20],
                        help="Sizes of problem instances (default 20, 50, 100)")
    parser.add_argument('--seed', type=int, default=2020, help="Random seed")

    args = parser.parse_args()
    np.random.seed(args.seed)

    try:
        for i in args.instance_sizes:
            dirs = ["./{}/{}".format(args.data_dir, i)]
            filenames = ["./{}/{}/{}_{}.xls".format(args.data_dir, i, args.filename, counter) for counter in
                         range(1, args.dataset_size + 1)]
            for dir in dirs:
                if not os.path.isdir(dir):
                    os.makedirs(dir)

            data = generate_cvrptw_data(args.dataset_size, i)
            write2xlsx(data, filenames)
    except Exception as e:
        print("---- [X] Generating Dataset Failed.")
        print(e)
    else:
        print("---- [V] Generating Dataset Successfully.")
