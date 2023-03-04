# import os, sys
# if 'SUMO_HOME' in os.environ:
#     tools = os.path.join(os.environ['SUMO_HOME'], 'tools')
#     sys.path.append(tools)
# else:
#     sys.exit("please declare environment variable 'SUMO_HOME'")

import openpyxl
import sumolib
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
from collections import namedtuple

def load_config_data(operation_path = 'data/config/operation_W.xlsx',
        net_path = 'data/config/airport.net.xml', cfg_path = 'data/config/testPrj.cfg'):
    operation = load_operation(operation_path) # fleets, services
    map = Map(net_path, cfg_path)

    return operation, map

def load_operation(path):
    data = excel_reader(path)
    # named tuple that will store info of a fleet (of a task) or a service (of a task&type)
    FleetInfo = namedtuple("FleetInfo", "name num capacity speed")
    ServiceInfo = namedtuple("ServiceInfo", "precedence_before precedence_after resource duration")

    task_names = data["Vehicles"][0][1:] # Read types
    tasks_id = list(enumerate(task_names, 1)) # assign id to tasks, starting from 1
    tasks = [x[0] for x in tasks_id]
    types = [int(i) for i in list(set(data["Operations"][0][1:]))] # Read types
    # Load fleet info - Sheet 3
    fleets = {task: FleetInfo(name = task_name, num = int(data["Vehicles"][1][task]), \
            capacity = int(data["Vehicles"][2][task]), speed = int(data["Vehicles"][3][task])/60 ) for task, task_name in tasks_id}

    # Load service info
    get_id_from_name = {name: id for id, name in tasks_id}

    # Sheet 2 - precedence
    precedence_before_dict = {}
    precedence_after_dict = {}
    for type in types:
        precedence_before_dict[type] = {}
        precedence_after_dict[type] = {}
        for task in tasks:
            precedence_before_dict[type][task] = []
            precedence_after_dict[type][task] = []

    for i in range(1, len(data["Precedences"][0])):

        type = int(data["Precedences"][0][i])
        from_task = get_id_from_name[data["Precedences"][1][i]]
        to_task = get_id_from_name[data["Precedences"][2][i]]

        precedence_before_dict[type][to_task].append(from_task)
        precedence_after_dict[type][from_task].append(to_task)

    # Sheet 1 - resource and duration
    duration_dict = {}
    resource_dict = {}
    for type in types:
        duration_dict[type] = {}
        resource_dict[type] = {}
    for i in range(1, len(data["Operations"][0])):
        type = int(data["Operations"][0][i])
        task = get_id_from_name[data["Operations"][1][i]]
        duration = int(data["Operations"][2][i]) * 60 # Unit: second
        resource = int(data["Operations"][3][i])
        # Add into dictionary
        duration_dict[type][task] = duration
        resource_dict[type][task] = resource

    # Store in named tuple
    services = {type: {task: ServiceInfo(
            precedence_before = precedence_before_dict[type][task],
            precedence_after = precedence_after_dict[type][task],
            resource = resource_dict[type][task],
            duration = duration_dict[type][task]) for task in tasks} for type in types}

    return {"fleets": fleets, "services": services, "tasks": tasks, "types": types}

def load_flights(path):
    data = excel_reader(path)
    sheet = data["Sheet1"]
    # Name tuple that stores flight info
    FlightInfo = namedtuple("FlightInfo", "flight_no arrival departure location type")
    flights = {id: FlightInfo(
            flight_no = sheet[1][id],
            arrival = (lambda x: x.hour*3600 + x.minute*60 + x.second)(sheet[2][id]),
            departure = (lambda x: x.hour*3600 + x.minute*60 + x.second)(sheet[3][id]),
            location = sheet[4][id],
            type = int(sheet[5][id])) for id in range(1, len(sheet[0]))
            }
    return flights

class Map:
    def __init__(self, netFilePath, prjCfgPath):
        self.net = sumolib.net.readNet(netFilePath)  # airport.net.xml
        self.data = self.__load_location(prjCfgPath)

    def load_distance(self, locations):
        # load distance between flights
        loc_list = locations
        distance =  {(i, j):
                self.net.getShortestPath(
                self.net.getEdge(self.__findedge(i, self.data)[1]), self.net.getEdge(self.__findedge(j, self.data)[0]))[1]
                for i in loc_list for j in loc_list }
        for i in loc_list:
            distance[(i, i)] = 0
        return distance

    def __load_location(self, prjXMLPath):
        tree = ET.parse(prjXMLPath)
        root = tree.getroot()

        # Read all locations and store as a list of ditct
        locationList = []
        for loc in root.find('locations').findall('loc'):
            locationList.append(loc.attrib)
        return locationList

    def __findedge(self, gate, data):
        edge_list = []
        for i in range(len(data)):
            if data[i]['name'] == gate:
                edge_list.append(data[i]['inEdgeID'])
                edge_list.append(data[i]['outEdgeID'])
                break
        return edge_list

def excel_reader(name):
    data = {}
    wb = openpyxl.load_workbook(name)
    for sheet in wb:
        table = [[cell.value for cell in col] for col in sheet.iter_cols()]
        data[sheet.title] = table
    return data

def excel_writer(data, path):
    # data: {sheet_name: tuple of tuple(rows)}
    wb = openpyxl.Workbook()
    for sheet_name in data:
        ws = wb.create_sheet(sheet_name)
        for row in data[sheet_name]:
            ws.append(row)
    del wb["Sheet"]
    wb.save(path)
    
def get_tuple(data):
    # Calculate average and convert nested list to nested tuple
    for key in data:
        # Calculate average
        sheet = data[key]
        average = [sum([sheet[row][column] for row in range(1, len(sheet))])/ (len(sheet)-1)
                for column in range(1, len(sheet[0]))]
        sheet.append(["Average"])
        sheet[-1].extend(average)
        # Change to nested tuple
        data[key] = tuple([tuple(row) for row in sheet])
