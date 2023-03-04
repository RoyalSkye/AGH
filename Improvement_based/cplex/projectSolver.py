#!/usr/bin/env python

import os, sys, xlrd
import numpy as np
from datetime import datetime, timedelta
if 'SUMO_HOME' in os.environ:
    tools = os.path.join(os.environ['SUMO_HOME'], 'tools')
    sys.path.append(tools)
else:
    sys.exit("please declare environment variable 'SUMO_HOME'")
import sumolib
import xml.etree.ElementTree as ET
import math
import time as Time

import xlwt
from openpyxl.workbook import Workbook as openpyxlWorkbook

import pickle
from alns_cvrptw import *
from utils import *


class Schedule_opr():

    def __init__(self):

        self.book = xlwt.Workbook()

    def sheet(self, lt1, lt2, lt3, lt4, lt5, lt6, lt7, lt8, lt9, lt10):

        sh = self.book.add_sheet('Sheet1')

        col1_name = 'Item No.'
        col2_name = 'Flight No.'
        col3_name = 'Operation'
        col4_name = 'Start Time'
        col5_name = 'End Time'
        col6_name = 'Vehicle Name'
        col7_name = 'From Gate'
        col8_name = 'Departure Time'
        col9_name = 'Arrival Time'
        col10_name = 'To Gate'
        # col11_name = 'Predict Delay'

        n = 0
        sh.write(n, 0, col1_name)
        sh.write(n, 1, col2_name)
        sh.write(n, 2, col3_name)
        sh.write(n, 3, col4_name)
        sh.write(n, 4, col5_name)
        sh.write(n, 5, col6_name)
        sh.write(n, 6, col7_name)
        sh.write(n, 7, col8_name)
        sh.write(n, 8, col9_name)
        sh.write(n, 9, col10_name)
        # sh.write(n, 10, col11_name)

        for m, e1 in enumerate(lt1, n + 1):
            sh.write(m, 0, e1)
        for m, e2 in enumerate(lt2, n + 1):
            sh.write(m, 1, e2)
        for m, e3 in enumerate(lt3, n + 1):
            sh.write(m, 2, e3)
        for m, e4 in enumerate(lt4, n + 1):
            sh.write(m, 3, e4)
        for m, e5 in enumerate(lt5, n + 1):
            sh.write(m, 4, e5)
        for m, e6 in enumerate(lt6, n + 1):
            sh.write(m, 5, e6)
        for m, e7 in enumerate(lt7, n + 1):
            sh.write(m, 6, e7)
        for m, e8 in enumerate(lt8, n + 1):
            sh.write(m, 7, e8)
        for m, e9 in enumerate(lt9, n + 1):
            sh.write(m, 8, e9)
        for m, e10 in enumerate(lt10, n + 1):
            sh.write(m, 9, e10)
        # for m, e11 in enumerate(lt11, n + 1):
        #     sh.write(m, 10, e11)

        return self.book


def excelreader(name):
    dic = {}
    workbook = xlrd.open_workbook(name)
    sheet_count = workbook.nsheets
    for i in range(sheet_count):
        worksheet = workbook.sheet_by_index(i)
        total_rows = worksheet.nrows
        total_cols = worksheet.ncols
        table = list()
        attribute = list()
        for y in range(total_cols):
            for x in range(total_rows):
                attribute.append(worksheet.cell(x, y).value)
            table.append(attribute)
            attribute = []
            x += 1
        dic[i] = table
    return dic, total_rows


def from_excel_ordinal(ordinal, _epoch0=datetime(1899, 12, 31)):
    if ordinal > 59:
        ordinal -= 1  # Excel leap year bug, 1900 is not a leap year!
    return (_epoch0 + timedelta(days=ordinal)).replace(microsecond=0)


def Loadlocation(prjXMLPath):
    prjFolder = os.path.dirname(prjXMLPath)
    tree = ET.parse(prjXMLPath)
    root = tree.getroot()

    # Read all locations and store as a list of ditct
    locationList = []
    for loc in root.find('locations').findall('loc'):
        locationList.append(loc.attrib)
    return locationList


def findedge(gate, data):
    edge_list = []
    for i in range(len(data)):
        if data[i]['name'] == gate:
            edge_list.append(data[i]['inEdgeID'])
            edge_list.append(data[i]['outEdgeID'])
            break
    return edge_list


# Read info from the input files, construct the ILP model,
# and solve it using CPLEX. The result file will be
# generated using the resultFilePath.
def solveProject(prjCfgPath, schFilePath, opnFilePath, netFilePath, resultFilePath, learning=True, evaluate=False):
    """
    learning: used to collect dataset, it will run 4 LNS with different heuristics and choose the best one.
    evaluate: evaluate model performance, it will run LNS based on model prediction.
    """
    xx, xy = excelreader(schFilePath)  # schedule_W_35_original.xlsx
    yx, yy = excelreader(opnFilePath)  # operation_W.xlsx

    # customer number
    n = len(xx[0][0]) - 1
    customers = [x for x in range(1, 1 + n)]

    rnd = np.random
    rnd.seed(1)

    #  fleet definition

    F = 10  # fleet number
    fleets = [x for x in range(1, 1 + F)]

    # add depot
    nodes = customers
    for i in range(F):
        nodes = nodes + [n + 1 + 2 * i]
        nodes = nodes + [n + 2 + 2 * i]

    nodes_F = {}
    for i in range(F):
        nodes_F[i + 1] = customers + [n + 1 + 2 * i] + [n + 2 + 2 * i]
    # print(nodes)
    # print(nodes_F)

    type_in_schedule = xx[0][5][1:len(xx[0][5])]
    type_in_schedule = [int(x) for x in type_in_schedule]
    # flight_type for extracting x's feature
    flight_type = {i: j for i, j in enumerate(type_in_schedule, start=1)}
    for i in range(F):
        flight_type[n + 1 + 2 * i] = 4
        flight_type[n + 2 + 2 * i] = 4
    demand = yx[0][3][1:len(yx[0][3])]
    demand = [int(x) for x in demand]

    D = {}
    for i in fleets:
        li = list()
        for j in type_in_schedule:
            li.append(demand[10 * (j - 1) + (i - 1)])
        D[i] = li

    Q_of_vehicle = yx[2][2][1:len(yx[2][1])]
    Q_of_vehicle = [int(x) for x in Q_of_vehicle]

    # read the distance from
    net = sumolib.net.readNet(netFilePath)  # airport.net.xml

    ret = xx[0][4][1:len(xx[0][4])] + ['depot'] * 2 * F

    distance_ret = dict(zip(nodes, ret))

    data = Loadlocation(prjCfgPath)

    distance = {
        (i, j): net.getShortestPath(
            net.getEdge(findedge(distance_ret[i], data)[1]), net.getEdge(findedge(distance_ret[j], data)[0]))[1]
        for i in nodes for j in nodes}

    for i in nodes:
        distance[i, i] = 0

    for i in range(F):
        distance[n + 1 + 2 * i, n + 2 + 2 * i] = 0
        distance[n + 2 + 2 * i, n + 1 + 2 * i] = 0

    # 05/15/2020 - Solve sumo problem, for 100 customers
    # with open('./test_200304/distance.pickle', 'rb') as f:
    #     distance = pickle.load(f)
    # 05/15/2020 END

    # Time windows

    # 10 ** 6 --> 10 ** 8 20191030
    I = 10 ** 8

    start_ = xx[0][2][1:len(xx[0][2])]
    end_ = xx[0][3][1:len(xx[0][3])]

    start_ea = {}
    for i in customers:
        string = start_[i - 1]
        tmpDatetime = from_excel_ordinal(string)  # Convert float to datetime
        time = [tmpDatetime.hour, tmpDatetime.minute]
        # minute --> second 20191030
        start_ea[i] = (time[0] * 60 + time[1]) * 60

    start_la = {}
    for i in customers:
        string = end_[i - 1]
        tmpDatetime = from_excel_ordinal(string)  # Convert float to datetime
        time = [tmpDatetime.hour, tmpDatetime.minute]
        # minute --> second 20191030
        start_la[i] = (time[0] * 60 + time[1]) * 60

    for i in range(2 * F):
        start_ea[n + i + 1] = -3600
        start_la[n + i + 1] = I

    duration = yx[0][2][1:len(yx[0][2])]
    q = {}
    for i in fleets:
        li = list()
        for j in type_in_schedule:
            # minute --> second 20191030
            li.append(60 * duration[10 * (j - 1) + (i - 1)])
        for k in range(2 * F):
            li.append(0)
        q[i] = li
    S_time = {j: q[j] for j in fleets}

    speed = yx[2][3][1:len(yx[2][3])]
    speed = [int(x) for x in speed]
    # minute --> second 20191030
    tau = {(i, j, f): S_time[f][i - 1] + 60 * distance[i, j] / (speed[f - 1])  # ADJUST SPEED HERE
           for i in nodes for j in nodes for f in fleets if i != j}

    # print(start_ea)

    delay_avg = xx[0][6][1:len(xx[0][6])]
    for i in range(2 * F):
        delay_avg.append(0)

    delay_var = xx[0][7][1:len(xx[0][7])]
    for i in range(2 * F):
        delay_var.append(0)

    # sample delay 20191230
    n_sample = 1000
    delay_sample = []
    for i in range(0, n):
        delay_sample.append(np.random.normal(loc=delay_avg[i], scale=delay_var[i], size=n_sample))
    delay_sample = np.array(delay_sample)
    delay_sample = delay_sample*60

    # TODO: 06/11/2020 Modified
    indices_tau = {(i, f) for f in fleets for i in nodes_F[f]}
    # initial solution
    t = {(i, f): 0 + start_ea[i] for i, f in indices_tau}
    # 1 calculate start time of each fleet according to precedence
    # remove int(max(...)) and add t[n+2*f-1,f]=0, t[n+2*f,f]=I 04/24/2020
    operation_type = yx[2][0][1:len(yx[2][0])]
    retrieve = dict(zip(operation_type, range(1, len(operation_type) + 1)))  # ["IB": 1, ...]
    ty = yx[1][0][1:len(yx[1][0])]
    ty = [int(x) for x in ty]
    for i in customers:
        Ty = type_in_schedule[i - 1]
        ind = [i for i, x in enumerate(ty) if x == Ty]
        for j in ind:
            t[i, retrieve[yx[1][2][1:len(yx[1][2])][j]]] = \
                max(t[i, retrieve[yx[1][2][1:len(yx[1][2])][j]]],
                    t[i, retrieve[yx[1][1][1:len(yx[1][1])][j]]] +
                    S_time[retrieve[yx[1][1][1:len(yx[1][1])][j]]][i - 1])
    for f in fleets:
        t[n + f * 2 - 1, f], t[n + 2 * f, f] = -3600, I

    # calculate the vehicle number according to time and capacity, generate the route at the same time
    # 根据TW 估计车辆的数量的时候 生成route 然后根据capacity的约束调整route (+vehicle) -> 根据precedence来决定哪个任务先执行t
    # 1 time window
    time_approximation = {f: [] for f in fleets}
    vehicle_number_approximation = {f: 0 for f in fleets}
    for f in fleets:
        for i in customers:
            for j in customers:
                if i != j:
                    time_approximation[f].append(tau[i, j, f])
    time_approximation_avg = {f: np.max(time_approximation[f]) for f in fleets}
    # print(' ')
    # print('period duration', time_approximation_avg)
    # print(' ')
    flight_start = list(start_ea.values())[0: n]

    time_vehicle = {f: {i: 0 for i in customers} for f in fleets}
    route = dict()
    for f in fleets:
        vehicle_count = 1
        while 0 in list(time_vehicle[f].values()):
            route[f, vehicle_count] = []
            current_time = 0
            max_time = max([t[j + 1, f] for j in range(n)])
            previous_flight = 0
            for j in range(n):
                if time_vehicle[f][j + 1] == 0:
                    current_time = t[j + 1, f] + S_time[f][j]
                    route[f, vehicle_count].append(j + 1)
                    time_vehicle[f][j + 1] = vehicle_count
                    previous_flight = j + 1
                    break
            nearest_flight = 1
            while nearest_flight != 0:
                nearest_distance = 10 ** 6
                nearest_flight = 0
                for j in range(n):
                    if time_vehicle[f][j + 1] == 0:
                        wait = t[j + 1, f] - current_time - 60 * distance[previous_flight, j + 1] / (speed[f - 1])
                        # if 0 < wait < 10 * 60:
                        if wait > 0 and nearest_flight == 0:
                            nearest_distance = 60 * distance[previous_flight, j + 1] / (speed[f - 1])
                            nearest_flight = j + 1
                        elif 0 < wait < 60 and nearest_distance > 60 * distance[previous_flight, j + 1] / (speed[f - 1]):
                            nearest_distance = 60 * distance[previous_flight, j + 1] / (speed[f - 1])
                            nearest_flight = j + 1
                if nearest_flight != 0:
                    current_time = t[nearest_flight, f] + S_time[f][nearest_flight - 1]
                    route[f, vehicle_count].append(nearest_flight)
                    time_vehicle[f][nearest_flight] = vehicle_count
                    previous_flight = nearest_flight

            vehicle_count += 1
            # for j in range(n):
            #     max_arc_cost = 0
            #     for i in range(j + 1, n):
            #         if time_vehicle[f][i + 1] == 0 and tau[j + 1, i + 1, f] > max_arc_cost:
            #             max_arc_cost = tau[j + 1, i + 1, f]
            #     # if flight_start[j] > current_time and time_vehicle[f][j + 1] == 0:
            #     if t[j + 1, f] > current_time and time_vehicle[f][j + 1] == 0:
            #         time_vehicle[f][j + 1] = vehicle_count
            #         # current_time = flight_start[j] + time_approximation_avg[f]
            #         current_time = t[j + 1, f] + time_approximation_avg[f]
            #         route[f, vehicle_count].append(j + 1)
            #         # current_time = t[j + 1, f] + max_arc_cost
            # vehicle_count += 1
        # for k in range(1, vehicle_count):
        #     for i in range(n):
        #         if time_vehicle[f][i + 1] == k:
        #             route[f, k].append(i + 1)
        vehicle_number_approximation[f] = np.max(list(time_vehicle[f].values()))
    print('vehicle 1', vehicle_number_approximation)
    # print('route', route)
    # 2 capacity
    for f in fleets:
        if Q_of_vehicle[f - 1] != 0:
            for k in range(1, vehicle_number_approximation[f] + 1):
                # print([demand[(type_in_schedule[i - 1] - 1) * F + f - 1] for i in route[f, k]])
                temp_demand = np.cumsum([demand[(type_in_schedule[i - 1] - 1) * F + f - 1] for i in route[f, k]])
                temp_route = route[f, k]
                temp_check = temp_demand < Q_of_vehicle[f - 1]
                false_ind = list(np.where(temp_check == False)[0])
                temp_k = k
                while False in temp_check:
                    route[f, temp_k] = temp_route[:min(false_ind)]
                    vehicle_number_approximation[f] += 1
                    route[f, vehicle_number_approximation[f]] = temp_route[min(false_ind):max(false_ind) + 1]
                    temp_k = vehicle_number_approximation[f]
                    temp_route = temp_route[min(false_ind):max(false_ind) + 1]
                    temp_demand = temp_demand - temp_demand[min(false_ind) - 1]
                    temp_demand = temp_demand[min(false_ind):max(false_ind) + 1]
                    temp_check = temp_demand < Q_of_vehicle[f - 1]
                    false_ind = list(np.where(temp_check == False)[0])

    print('vehicle 2', vehicle_number_approximation)
    # 08/03/2020:
    # TODO: have bugs, cvrptw.visit may contains invalid x, aka (k, f)
    # 20:  vehicle{1: 5, 2: 7, 3: 8, 4: 7, 5: 6, 6: 7, 7: 7, 8: 7, 9: 6, 10: 6}
    # 50:  vehicle{1: 5, 2: 8, 3: 9, 4: 8, 5: 9, 6: 9, 7: 10, 8: 10, 9: 10, 10: 6}
    # 100: vehicle{1: 6, 2: 8, 3: 13, 4: 9, 5: 16, 6: 11, 7: 16, 8: 20, 9: 22, 10: 6}
    # 200: vehicle{1: 6, 2: 9, 3: 22, 4: 10, 5: 29, 6: 10, 7: 23, 8: 38, 9: 40, 10: 6}
    # vehicle_tmp = vehicle_number_approximation
    # vehicle_number_approximation = {1: 15, 2: 15, 3: 15, 4: 10, 5: 10, 6: 10, 7: 10, 8: 10, 9: 10, 10: 10}
    # # vehicle_number_approximation = {1: 10, 2: 10, 3: 15, 4: 10, 5: 20, 6: 15, 7: 20, 8: 25, 9: 25, 10: 10}
    # for f in fleets:
    #     for k in set(range(1, vehicle_number_approximation[f]+1)) - set(range(1, vehicle_tmp[f]+1)):
    #         route[f, k] = []
    # print('vehicle 3', vehicle_number_approximation)

    vehicles = {j: [x for x in range(1, 1 + vehicle_number_approximation[j])] for j in fleets}
    Q = {i: np.repeat(Q_of_vehicle[i - 1], vehicle_number_approximation[i]) for i in fleets}
    indices = {(i, j, k, f) for f in fleets for i in nodes_F[f] for j in nodes_F[f] for k in vehicles[f] if i != j}

    x = {(i, j, k, f): 0 for i, j, k, f in indices}

    print(' ')
    for key in route.keys():
        route_temp = route[key]
        if route_temp:
            for i in range(len(route_temp) - 1):
                x[route_temp[i], route_temp[i + 1], key[1], key[0]] = 1
            x[n + 2 * key[0] - 1, route_temp[0], key[1], key[0]] = 1
            x[route_temp[len(route_temp) - 1], n + 2 * key[0], key[1], key[0]] = 1
        else:
            x[n + 2 * key[0] - 1, n + 2 * key[0], key[1], key[0]] = 1  # start_depot -> end_depot

    init_solution = {"x": x, "t": t, "route": route}
    # print("init_solution: {}".format(init_solution))
    # 06/11/2020 END

    # 04/01/2020 LNS
    operation_type = yx[2][0][1:len(yx[2][0])]
    # retrieve: operation: fleet_id {'IB': 1, 'UL/L': 2, 'CI': 3, 'DB': 4, ...}
    retrieve = dict(zip(operation_type, range(1, len(operation_type) + 1)))
    from_task = yx[1][1][1:len(yx[1][1])]
    to_task = yx[1][2][1:len(yx[1][2])]
    ty = yx[1][0][1:len(yx[1][0])]
    ty = [int(x) for x in ty]
    # {1:[[],[]]}
    precedence = {}
    for i, type1 in enumerate(type_in_schedule, start=1):
        index = [id for id, type_x in enumerate(ty) if type_x == type1]
        precedence_item = []
        for j in index:
            precedence_item.append([retrieve[from_task[j]], retrieve[to_task[j]]])
        precedence[i] = precedence_item

    args = {"customers": customers,
            "fleets": fleets,
            "start_early": start_ea,
            "start_late": start_la,
            "nodes": nodes,
            "nodes_fleet": nodes_F,
            "travel_time": tau,
            "degree_of_destruction": 0.25,
            "vehicles": vehicles,
            "distance": distance,
            "demand_operation": D,
            "capacity_fleet": Q,
            "duration": S_time,
            "precedence": precedence,
            "flight_type": flight_type}

    cvrptw = CVRPTW(args=args, init_solution=init_solution, learning=learning)
    cvrptw_best = copy.deepcopy(cvrptw)
    valid = cvrptw.is_valid()
    if not valid:
        print("[!] Init Solution was Rejected.")
        exit(0)
    print("Init Solution Accepted? {}".format(valid))
    print("Init Solution Obj: {}".format(cvrptw.obj))
    print("Init Solution used {} vehicles".format(used_vehicle(cvrptw)))
    # 04/01/2020 END

    # 05/10/2020 Load a solution from checkpoint
    # load_sol(cvrptw, filepath='./result/error_sol.json')
    # load_sol(cvrptw_best, filepath='/Users/skye/docs/cvprtw_json/BEST_50_True_0.1_0.5_150622.json')  # for Visualization
    # 05/10/2020 END

    if evaluate:
        # cvrptw.learning = False
        # obj, res_list, time_list = lns(cvrptw, cvrptw_best, mipgap=0.1, timelimit=120, degree_of_destruction=0.7, iteration=5, shuffle="RANDOM", accept_rate=1.01)
        # cvrptw.learning = True
        return cvrptw, cvrptw_best
    elif learning:
        best_obj, best_res = 10 ** 6, None
        for count in range(10):
            res = lns(copy.deepcopy(cvrptw), copy.deepcopy(cvrptw_best), mipgap=0.1, timelimit=120, degree_of_destruction=0.7, iteration=1, shuffle="RANDOM", accept_rate=1.01)
            best_res = res if res[-1] < best_obj else best_res
            best_obj = min(best_obj, res[-1])
        return best_res
    else:
        res = []
        # 0. Random_LNS
        res.append(lns(copy.deepcopy(cvrptw), copy.deepcopy(cvrptw_best), mipgap=0.1, timelimit=60, degree_of_destruction=0.8, iteration=30, shuffle="RANDOM_LNS", accept_rate=1.01))
        # 1. Vehicle_Random
        res.append(lns(copy.deepcopy(cvrptw), copy.deepcopy(cvrptw_best), mipgap=0.1, timelimit=60, degree_of_destruction=0.7, iteration=30, shuffle="RANDOM", accept_rate=1.01))
        # 2. Vehicle_Descend
        res.append(lns(copy.deepcopy(cvrptw), copy.deepcopy(cvrptw_best), mipgap=0.1, timelimit=60, degree_of_destruction=0.45, iteration=30, shuffle="DISTANCE_DESCEND", accept_rate=1.01))
        # res.append(lns(copy.deepcopy(cvrptw), copy.deepcopy(cvrptw_best), mipgap=0.1, timelimit=120, degree_of_destruction=0.4, iteration=30, shuffle="DISTANCE_ASCEND", accept_rate=1.01))
        # 3. Vehicle_Descend_Random
        res.append(lns(copy.deepcopy(cvrptw), copy.deepcopy(cvrptw_best), mipgap=0.1, timelimit=60, degree_of_destruction=0.5, iteration=30, shuffle="DISTANCE_DESCEND", randomness=0.6, disjoint=False, accept_rate=1.01))
        return res

    # 05/06/2020
    # # Optimization
    # from docplex.mp.model import Model
    # mdl = Model('CVRP_multi_fleet')
    # # adjust 100 flights parameters 20191119
    # # mdl.parameters.mip.strategy.file = 3  # compress the nodes and store them on disk
    # # mdl.parameters.emphasis.mip = 1  # 1 feasibility 2 optimality
    # # mdl.parameters.mip.strategy.fpheur = 1  # Turns on the feasibility pump heuristic for (MIP) models
    # # mdl.parameters.mip.strategy.variableselect = 3  # strong branching: a smaller list of active nodes
    # # mdl.parameters.mip.strategy.nodeselect = 0  # Depth-first search
    # # mdl.parameters.threads = 4
    # # mdl.parameters.preprocessing.symmetry = 5
    # # mdl.parameters.mip.strategy.kappastats = 1
    # 05/06/2020 END

    '''
    # Create the result file
    min_time = 18000
    sec_min = min_time % 60
    miu_min = min_time // 60
    hour_min = miu_min // 60
    miu_min = miu_min % 60
    str_min_time = str(int(hour_min)) + ':' + str(int(miu_min)) + ':' + str(int(sec_min))
    time_ret = {}
    for k in indices_tau:
        # 05/13/2020
        # time_ret[k] = t[k].solution_value
        time_ret[k] = cvrptw_best.time[k]
        # 05/13/2020 END

    lt1 = [x for x in range(1, len(customers) * 10 + 1)]
    item_count = len(customers) * 10 + 1

    lt2 = []
    for i in range(1, len(xx[0][1])):
        for j in range(10):
            lt2.append(xx[0][1][i])

    lt3 = []
    for i in range(len(customers)):
        for j in range(10):
            lt3.append(operation_type[j])

    lt4 = []
    lt5 = []
    lt6 = []
    lt7 = []
    lt8 = []
    lt9 = []
    lt10 = []
    # lt11 = []

    # f_delay = open('Delay.pickle', 'wb')
    # flight_delay = {}

    for i in range(1, len(customers) + 1):
        # pre_delay = 'Predict Delay: ' + str(round(predict_delay[i].solution_value/60, 2)) + ' minutes'
        # flight_delay[i-1] = pre_delay
        for f in range(1, 11):
            # 05/13/2020
            # ind = [k for k in indices if x[k].solution_value > 0.9 and k[3] == f and k[1] == i]
            ind = [k for k in indices if cvrptw_best.visit[k] > 0.9 and k[3] == f and k[1] == i]
            # 05/13/2020 END
            fro = ind[0][0]
            to = ind[0][1]
            vn = ind[0][2]
            fln = ind[0][3]
            sta = time_ret[(to, fln)]
            # S_time[fln][i-1] --> S_time[fln][to-1]
            end = time_ret[(to, fln)] + S_time[fln][to - 1]
            # add second 20191030
            sec_s = sta % 60
            miu_s = sta // 60
            hour_s = miu_s // 60
            miu_s = miu_s % 60
            sec_e = end % 60
            miu_e = end // 60
            hour_e = miu_e // 60
            miu_e = miu_e % 60
            stri_s = str(int(hour_s)) + ':' + str(int(miu_s)) + ':' + str(int(sec_s))
            stri_e = str(int(hour_e)) + ':' + str(int(miu_e)) + ':' + str(int(sec_e))
            lt4.append(stri_s)
            lt5.append(stri_e)
            lt6.append(operation_type[f - 1] + '(' + str(vn) + ')')
            # ret[fro] --> ret[fro-1]
            lt7.append(ret[fro - 1])  # from
            lt10.append(ret[to - 1])

            # S_time[fln][i-1] --> S_time[fln][fro-1]
            if fro > n:
                dep = time_ret[to, fln] - 60 * distance[fro, to] / speed[fln - 1]
            else:
                dep = time_ret[fro, fln] + S_time[fln][fro - 1]
            # add second 20191030
            sec_d = dep % 60
            miu_d = dep // 60
            hour_d = miu_d // 60
            miu_d = miu_d % 60
            stri_d = str(int(hour_d)) + ':' + str(int(miu_d)) + ':' + str(int(sec_d))
            lt8.append(stri_d)

            # change speed here
            arr = dep + 60 * distance[fro, to] / speed[fln - 1]  # /3.6)
            # add second 20191030
            sec_a = arr % 60
            miu_a = arr // 60
            hour_a = miu_a // 60
            miu_a = miu_a % 60
            stri_a = str(int(hour_a)) + ':' + str(int(miu_a)) + ':' + str(int(sec_a))
            lt9.append(stri_a)

            # lt11.append(predict_delay[i].solution_value/60)

    # pickle.dump(flight_delay, f_delay)
    # f_delay.close()

    lt6_temp = [a for a in range(len(lt6))]
    lt6_forSearch = dict(zip(lt6, lt6_temp))

    f_utilization = open('Utilization.pickle', 'wb')
    fleet_utilization = {}

    for f in fleets:
        utilization_count = len(vehicles[f])
        for v in vehicles[f]:
            vehicle_temp = operation_type[f - 1] + '(' + str(v) + ')'
            if vehicle_temp not in lt6_forSearch:
                utilization_count = utilization_count - 1
                lt1.append(item_count)
                item_count = item_count + 1
                lt2.append(operation_type[f - 1] + '_Depot')
                lt3.append(operation_type[f - 1])
                lt4.append(str_min_time)
                lt5.append(str_min_time)
                lt6.append(operation_type[f - 1] + '(' + str(v) + ')')
                lt7.append('depot')
                lt8.append(str_min_time)
                lt9.append(str_min_time)
                lt10.append('depot')
                # lt11.append(0)
        print(str(operation_type[f - 1]) + ': ' +
              str(utilization_count) + ' / ' + str(len(vehicles[f])) +
              ' = ' + str(100 * utilization_count / len(vehicles[f])) + ' %')
        utilization = 'Utilization: ' + str(utilization_count) + ' / ' + \
                      str(len(vehicles[f])) + ' = ' + str(100 * utilization_count / len(vehicles[f])) + ' %'
        fleet_utilization[operation_type[f-1]] = utilization
    pickle.dump(fleet_utilization, f_utilization)

    f_utilization.close()

    # XLS
    op = Schedule_opr()
    a = op.sheet(lt1, lt2, lt3, lt4, lt5, lt6, lt7, lt8, lt9, lt10)
    a.save(resultFilePath)

    # conversion to XLSX
    xlsBook = xlrd.open_workbook(resultFilePath)
    workbook = openpyxlWorkbook()

    for i in range(0, xlsBook.nsheets):
        xlsSheet = xlsBook.sheet_by_index(i)
        sheet = workbook.active if i == 0 else workbook.create_sheet()
        sheet.title = xlsSheet.name

        for row in range(0, xlsSheet.nrows):
            for col in range(0, xlsSheet.ncols):
                if row > 0 and (col == 3 or col == 4 or col == 7 or col == 8):
                    stri = xlsSheet.cell_value(row, col)
                    # time = [int(s) for s in stri.split(":") if s.isdigit()]
                    # datetime_object = datetime.strptime(stri, "%H:%M")
                    # add second 20191030
                    datetime_object = datetime.strptime(stri, "%H:%M:%S")
                    # datetime_object = datetime_object.strftime("%H:%M:%S")
                    sheet.cell(row=row + 1, column=col + 1).value = datetime_object - datetime(1900, 1, 1)
                    # sheet.cell(row=row + 1, column=col + 1).number_format
                else:
                    sheet.cell(row=row + 1, column=col + 1).value = xlsSheet.cell_value(row, col)

    workbook.save(resultFilePath)
    '''


def extract_k_f_feature(cvrptw, sol=True):
    """
    extract k_f's feature: (k, f): 0/1 whether used or not + distance(k, f)-normalized + f 10-dim one-hot to distinguish
    """
    one_hot = {1: [1, 0, 0, 0, 0, 0, 0, 0, 0, 0], 2: [0, 1, 0, 0, 0, 0, 0, 0, 0, 0], 3: [0, 0, 1, 0, 0, 0, 0, 0, 0, 0],
               4: [0, 0, 0, 1, 0, 0, 0, 0, 0, 0], 5: [0, 0, 0, 0, 1, 0, 0, 0, 0, 0], 6: [0, 0, 0, 0, 0, 1, 0, 0, 0, 0],
               7: [0, 0, 0, 0, 0, 0, 1, 0, 0, 0], 8: [0, 0, 0, 0, 0, 0, 0, 1, 0, 0], 9: [0, 0, 0, 0, 0, 0, 0, 0, 1, 0], 10: [0, 0, 0, 0, 0, 0, 0, 0, 0, 1]}
    k_f = [0] * len(cvrptw.veh_map)
    vehicle_distance = {(k, f): 0 for f in cvrptw.args["fleets"] for k in cvrptw.args["vehicles"][f]}
    for key, value in cvrptw.visit.items():
        if value != 1:
            continue
        i, j, k, f = key[0], key[1], key[2], key[3]
        vehicle_distance[(k, f)] += cvrptw.args["distance"][(i, j)]

    max_value = max(vehicle_distance.values())

    if sol:
        for key, value in cvrptw.veh_map.items():
            is_visited = 0 if vehicle_distance[key] == 0 else 1
            k_f[value] = [is_visited, round(vehicle_distance[key]/max_value, 4)]
            # k_f[value].extend(one_hot[key[-1]])
    else:
        for key, value in cvrptw.veh_map.items():
            k_f[value] = [0, 0]
            # k_f[value].extend(one_hot[key[-1]])
    cvrptw.data["k_f"].append(k_f)


def extract_v_feature(cvrptw, cvrptw_best, sol=True):
    """
    extract v's feature:
    x: obj_coefficients-normalized + sol_val + inc_val + type(i) one-hot + type(j) one-hot + duration(i)-normalized + duration(j)-normalized + t[i, f]-normalized+ t[j, f]-normalized
    t: sol_val-normalized, inc_val-normalized
    """
    one_hot = {1: [1, 0, 0, 0], 2: [0, 1, 0, 0], 3: [0, 0, 1, 0], 4: [0, 0, 0, 1]}  # type 4: depot
    x_vec = [0] * len(cvrptw.visit)
    t_vec = [0] * len(cvrptw.time)
    max_distance = max(cvrptw.args["distance"].values())
    max_duration, max_sol_t, max_inc_t = 0, 86400, 86400
    for i in cvrptw.args["duration"].values():
        max_duration = max(max_duration, max(i))
    # v_vec = [0] * (len(cvrptw.visit) + len(cvrptw.time))
    if sol:
        for k in cvrptw.visit.keys():
            t1 = round(cvrptw.time[k[0], k[-1]]/max_sol_t, 4) if cvrptw.time[k[0], k[-1]] >= 0 and cvrptw.time[k[0], k[-1]] < 1000000 else 0
            t2 = round(cvrptw.time[k[1], k[-1]]/max_sol_t, 4) if cvrptw.time[k[1], k[-1]] >= 0 and cvrptw.time[k[1], k[-1]] < 1000000 else 0
            x_vec[cvrptw.id_map[k[0], k[1], k[2], k[3]]] = [round(cvrptw.args["distance"][k[0], k[1]]/max_distance, 4), cvrptw.visit[k], cvrptw_best.visit[k]]
            # x_vec[cvrptw.id_map[k[0], k[1], k[2], k[3]]].extend(one_hot[cvrptw.args["flight_type"][k[0]]])
            # x_vec[cvrptw.id_map[k[0], k[1], k[2], k[3]]].extend(one_hot[cvrptw.args["flight_type"][k[1]]])
            # x_vec[cvrptw.id_map[k[0], k[1], k[2], k[3]]].extend([round(cvrptw.args["duration"][k[-1]][k[0]-1]/max_duration, 4), round(cvrptw.args["duration"][k[-1]][k[1]-1]/max_duration, 4), t1, t2])
        for k in cvrptw.time.keys():
            t1 = round(cvrptw.time[k]/max_sol_t, 4) if cvrptw.time[k] >= 0 and cvrptw.time[k] < 1000000 else 0
            t2 = round(cvrptw_best.time[k]/max_inc_t, 4) if cvrptw_best.time[k] >= 0 and cvrptw_best.time[k] < 1000000 else 0
            t_vec[cvrptw.id_map[k[0], k[-1]] - len(cvrptw.visit)] = [t1, t2]
    else:
        for k in cvrptw.visit.keys():
            x_vec[cvrptw.id_map[k[0], k[1], k[2], k[3]]] = [round(cvrptw.args["distance"][k[0], k[1]]/max_distance, 4), 0, cvrptw_best.visit[k]]
            # x_vec[cvrptw.id_map[k[0], k[1], k[2], k[3]]].extend(one_hot[cvrptw.args["flight_type"][k[0]]])
            # x_vec[cvrptw.id_map[k[0], k[1], k[2], k[3]]].extend(one_hot[cvrptw.args["flight_type"][k[1]]])
            # x_vec[cvrptw.id_map[k[0], k[1], k[2], k[3]]].extend([round(cvrptw.args["duration"][k[-1]][k[0]-1]/max_duration, 4), round(cvrptw.args["duration"][k[-1]][k[1]-1]/max_duration, 4), 0, 0])
        for k in cvrptw.time.keys():
            t2 = round(cvrptw_best.time[k]/max_inc_t, 4) if cvrptw_best.time[k] >= 0 and cvrptw_best.time[k] < 1000000 else 0
            t_vec[cvrptw.id_map[k[0], k[-1]] - len(cvrptw.visit)] = [0, t2]
    # cvrptw.data["v"].append(v_vec)
    cvrptw.data["x"].append(x_vec)
    cvrptw.data["t"].append(t_vec)


def lns(cvrptw, cvrptw_best, mipgap=0.1, timelimit=60, degree_of_destruction=0.7, iteration=30, shuffle="RANDOM", randomness=None, disjoint=False, accept_rate=1.01):
    # 05/06/2020 destroy and repair
    # TODO: Accept: SAA.
    print("LNS BEGIN")
    seed_everything(seed=2020)
    start_time = Time.time()
    start, reject = 0, 0
    randomness = randomness
    disjoint = disjoint
    shuffle = shuffle  # NONE, RANDOM, DISTANCE_DESCEND, DISTANCE_ASCEND, BOTH_ENDS
    res_list = [cvrptw.obj]
    time_list = [0]
    iteration = iteration
    mipgap = mipgap
    timelimit = timelimit
    accept_rate = accept_rate
    degree_of_destruction = degree_of_destruction

    # 06/27/2020 BEGIN
    # For by_vehicle_destroy
    cvrptw.vehicle_fleet = [(k, f) for f in cvrptw.args["fleets"] for k in cvrptw.args["vehicles"][f]]
    shuffle_vehicle_fleet(cvrptw, shuffle=shuffle)
    # 06/27/2020 END

    # for initial solution
    if cvrptw.learning:
        extract_v_feature(cvrptw, cvrptw_best, sol=True)
        extract_k_f_feature(cvrptw, sol=True)

    # For Random_LNS and Vehicle_LNS
    for i in range(iteration):
        if randomness is None:
            if shuffle == "RANDOM_LNS":
                solution = random_destroy(cvrptw, degree_of_destruction=degree_of_destruction, mipgap=mipgap, timelimit=timelimit)
            else:
                solution, start = by_vehicle_destroy(cvrptw, degree_of_destruction=degree_of_destruction, mipgap=mipgap, timelimit=timelimit, start=start, shuffle=shuffle)
                # solution = by_fleet_destroy(cvrptw, fleet=i, mipgap=mipgap, timelimit=timelimit)
        else:
            solution, start = by_vehicle_destroy_randomness(cvrptw, degree_of_destruction=degree_of_destruction, randomness=randomness,
                                                mipgap=mipgap, timelimit=timelimit, start=start, shuffle=shuffle, disjoint=disjoint)

        time_list.append((int(Time.time() - start_time)))
        if solution is None:
            print("{} - Cplex doesn't find a solution within {}s".format(i + 1, timelimit))
            res_list.append(None)
            reject = reject + 1
            print("Time {}".format(time_list[-1]))
            cvrptw.accepted.append(False)
            if cvrptw.learning:
                if i != iteration-1:
                    extract_v_feature(cvrptw, cvrptw_best, sol=False)
                    extract_k_f_feature(cvrptw, sol=False)
                else:
                    cvrptw.data["accepted"] = cvrptw.accepted
            continue
        # evaluate and accept
        new_sol_obj = int(solution.objective_value)
        res_list.append(new_sol_obj)
        print("Old -> New Solution objective: {} -> {}".format(cvrptw.obj, new_sol_obj))
        if new_sol_obj < cvrptw.obj * accept_rate:
            reject = 0
            cvrptw.accepted.append(True)
            visit = {k: int(cvrptw.x[k].solution_value) for k in cvrptw.visit.keys()}
            t1 = {k: cvrptw.t[k].solution_value for k in cvrptw.time.keys()}
            cvrptw.set_solution(visit, t1)
            if not cvrptw.is_valid():
                save_sol(visit, t1, filepath="./error_sol.json")
                print("New solution doesn't pass constraint test!")
                return
            print("{} - New Solution is accepted when accept_rate is {}, used {} vehicles".format(i + 1, accept_rate,
                                                                                                  used_vehicle(cvrptw)))
            # substitute current solution
            if new_sol_obj < cvrptw_best.obj:
                cvrptw_best.set_solution(visit, t1)
                # save_sol(visit, t1)
        else:
            reject = reject + 1
            cvrptw.accepted.append(False)
            if reject >= 2:
                reject = 0
                # accept_rate = accept_rate + 0.01
                mipgap = mipgap * 0.9  # 0.85
                degree_of_destruction = degree_of_destruction * 0.97
                # print("Increase accept_rate to {}".format(accept_rate))
                print("Reduce mipgap to {}".format(mipgap))
                print("Reduce degree_of_destruction to {}".format(degree_of_destruction))
            print("{} - New Solution is rejected.".format(i + 1))

        if cvrptw.learning:
            if i != iteration - 1:
                extract_v_feature(cvrptw, cvrptw_best, sol=True)
                extract_k_f_feature(cvrptw, sol=True)
            else:
                cvrptw.data["accepted"] = cvrptw.accepted

        print("Time {}".format(time_list[-1]))

    # for i in range(iteration):
    #     solution = local_branching(cvrptw, mipgap=mipgap, timelimit=timelimit, k=15)
    #     time_list.append((int(Time.time() - start_time)))
    #     if solution is None:
    #         print("{} - Cplex doesn't find a solution within {}s".format(i + 1, timelimit))
    #         res_list.append(None)
    #         print("Time {}".format(time_list[-1]))
    #         continue
    #     # evaluate and accept
    #     new_sol_obj = int(solution.objective_value)
    #     res_list.append(new_sol_obj)
    #     print("Old -> New Solution objective: {} -> {}".format(cvrptw.obj, new_sol_obj))
    #     if new_sol_obj < cvrptw.obj * accept_rate:
    #         visit = {k: int(cvrptw.x[k].solution_value) for k in cvrptw.visit.keys()}
    #         t1 = {k: cvrptw.t[k].solution_value for k in cvrptw.time.keys()}
    #         cvrptw.set_solution(visit, t1)
    #         if not cvrptw.is_valid():
    #             save_sol(visit, t1, filepath="./error_sol.json")
    #             print("New solution doesn't pass constraint test!")
    #             return
    #         print("{} - New Solution is accepted when accept_rate is {}, used {} vehicles".format(i + 1, accept_rate, used_vehicle(cvrptw)))
    #         # substitute current solution
    #         if new_sol_obj < cvrptw_best.obj:
    #             cvrptw_best.set_solution(visit, t1)
    #             save_sol(visit, t1)
    #     else:
    #         print("{} - New Solution is rejected.".format(i + 1))
    #     print("Time {}".format(time_list[-1]))

    print("AVG Construct Time {}s".format(cvrptw.construct_time / iteration))
    print("LNS END after {}s".format(int(Time.time() - start_time)))
    print("BEST_OBJ = {}".format(cvrptw_best.obj))
    print("BEST_SOL used {} Vehicles".format(used_vehicle(cvrptw_best)))
    save_sol(cvrptw_best.visit, cvrptw_best.time, filepath="./BEST_sol.json")
    print(res_list)
    print(time_list)
    for i, j in enumerate(res_list):
        print("{}: {}".format(i, j), end=" -> ")
    print(" ")
    # 05/06/2020 END

    if cvrptw.learning:
        return [cvrptw.id_map, cvrptw.veh_map, cvrptw.data, cvrptw_best.obj]
    else:
        return [cvrptw_best.obj, res_list, time_list]


if __name__ == "__main__":
    # TODO: To reproduce results, use seed_everything()
    data, id_map = {}, {}
    dir = "./data/Train1"
    prjCfgPath = "./test_200304/testPrj.cfg"
    schFilePath = "./test_200304/schedule_20.xlsx"
    opnFilePath = "./test_200304/operation_W.xlsx"
    netFilePath = "./test_200304/airport.net.xml"
    resultFilePath = "./test_200304/result.xlsx"

    # Normal LNS
    res = solveProject(prjCfgPath, schFilePath, opnFilePath, netFilePath, resultFilePath, learning=False, evaluate=False)
    exit(0)

    # 1. Imitation Learning
    # shared_data: some data that shared among different iterations in each instance
    # unshared_data: variable features that are different among iterations
    # file_map: instance_filename to id
    file_map, shared_data, unshared_data = {}, {}, {}
    files = os.listdir(dir)
    counter = 0
    for file in files:
        if os.path.splitext(file)[-1][1:] not in ["xlsx"]:
            print("Unsupported file detected!")
            continue
        path = os.path.join(dir, file)
        print(path)
        file_map[path] = counter
        id_map, veh_map, data, obj = solveProject(prjCfgPath, path, opnFilePath, netFilePath, resultFilePath, learning=True, evaluate=False)
        unshared_data[counter] = {"x": data["x"], "t": data["t"], "k_f": data["k_f"], "label": data["label"], "accepted": data["accepted"]}
        shared_data[counter] = {"obj": obj, "id_map": id_map, "veh_map": veh_map, "c": data["c"], "e": data["e"]}
        counter += 1

    create_datasets(shared_data, unshared_data, file_map, output_folder=dir)

    # file_map, unshared_data, shared_data = load_datasets(dir_path="./data")
    # assert len(file_map) == len(unshared_data) == len(shared_data)
    # assert len(unshared_data["0"]["v"]) == len(unshared_data["0"]["label"])
    # print(len(unshared_data) * len(unshared_data["0"]["label"]))
